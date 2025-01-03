import openpyxl


def get_data_fromEXEL(file_location):
    """
    method for reading sheets of an Excel file.

    Args:
        file_location (str): location of the target Excel file.

    Returns:
        list: [[list of all the test data rows], {dictionary of all the specific test values}, "final result"].

    """

    # getting a list of all the test data rows
    workbook = openpyxl.load_workbook(file_location)
    tests = workbook["tests"]
    all_rows = []
    for r in range(2, tests.max_row+1):
        row = []
        for c in range(1, tests.max_column+1):
            row.append(str(tests.cell(r,c).value))
        all_rows.append(row)

    # getting a dictionary of all the specific test values
    test_values = workbook["values"]
    values = {}
    for c in range(1, test_values.max_column+1):
        values[str(test_values.cell(1,c).value)] = str(test_values.cell(2,c).value)
    # specifying the value of "empty"
    values["empty"] = ""

    result = str(workbook["result"].cell(1,1).value)

    return [all_rows, values, result]
